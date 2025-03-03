# Importaciones
import openpyxl
from tabulate import tabulate # Ésta es una librería que toca descargar

# Leer el excel
texto = openpyxl.load_workbook("Practica 2/Datos vias.xlsx")
excel = texto.active

lista = []
# Recorrer el documento
for i in range(2, excel.max_row):
    fila = [] # Guardar cada fila en una lista
    for col in excel.iter_cols(0, excel.max_column):
        fila.append(col[i].value) # Pedimos el valor en la posición i
    lista.append(fila) # Agregamos a nuestra lista

# Mostramos la lista en una tablita (para verla más facil)
headers = ["ciudad 1","cuidad 2", "Kilometros", "minutos"]
headers_align = (("center",)*4)

print(tabulate(lista, headers=headers, tablefmt="fancy_grid" ,colalign=headers_align))